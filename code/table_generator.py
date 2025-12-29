"""
Класс TablesGenerator - генерация таблиц с результатами
"""

# Импорт необходимых для работы библиотек
import json
import numpy as np
import pandas as pd
from typing import List, Dict, Tuple
from openpyxl.styles import Alignment, Font

class TablesGenerator:
    """Класс для генерации таблиц в формате Excel"""
    def __init__(self, solver_results: Dict):
        """Инициализация с результатами расчета"""
        self.results = solver_results
        self.alternatives = solver_results['alternatives']
        self.criteria = solver_results['criteria']
        self.dms = solver_results['dms']
        self.dm_ids = solver_results['dm_ids']

        # Преобразование обратно в numpy массивы для удобства
        self.X_list = [np.array(X) for X in solver_results['X_list']]
        self.weights_list = [np.array(w) for w in solver_results['weights_list']]
        self.Y_list = [np.array(Y) for Y in solver_results['Y_list']]

        self.m = len(self.alternatives)
        self.n = len(self.criteria)
        self.t = len(self.dms)

    def normalize_matrix(self, X: np.ndarray) -> np.ndarray:
        """Нормализация матрицы решений (копия метода из Solver)"""
        R = np.zeros_like(X, dtype=float)

        for j in range(self.n):
            criterion_type = self.criteria[j]['type']
            column = X[:, j]

            if criterion_type == 'positive':
                # Benefit criterion
                norm = np.sqrt(np.sum(column**2))
                R[:, j] = column / norm if norm != 0 else 0
            elif criterion_type == 'negative':
                # Cost criterion
                norm = np.sqrt(np.sum(column**2))
                R[:, j] = 1 - (column / norm) if norm != 0 else 0
            else:
                raise ValueError(f"Unknown criterion type: {criterion_type}")

        return R

    def create_table_1(self) -> pd.DataFrame:
        """Создание Таблицы 1: Decision matrixes of example-subjective attributes"""
        table_data = []

        for i, alt in enumerate(self.alternatives):
            row = {'Candidate': alt}
            for k, dm in enumerate(self.dms):
                for j, criterion in enumerate(self.criteria):
                    col_name = f"{dm['id']}_{criterion['name']}"
                    row[col_name] = self.X_list[k][i, j]
            table_data.append(row)

        df = pd.DataFrame(table_data)
        return df

    def create_table_2(self) -> pd.DataFrame:
        """Создание Таблицы 2: Normalized decision matrixes"""
        table_data = []
        R_list = [self.normalize_matrix(X) for X in self.X_list]

        for i, alt in enumerate(self.alternatives):
            row = {'Candidate': alt}
            for k, dm in enumerate(self.dms):
                for j, criterion in enumerate(self.criteria):
                    col_name = f"{dm['id']}_{criterion['name']}"
                    row[col_name] = round(R_list[k][i, j], 4)
            table_data.append(row)

        df = pd.DataFrame(table_data)
        return df

    def create_table_3(self) -> pd.DataFrame:
        """Создание Таблицы 3: Weights on attributes of example"""
        table_data = []

        for j, criterion in enumerate(self.criteria):
            row = {'Criterion': criterion['name']}
            for k, dm in enumerate(self.dms):
                row[dm['id']] = round(self.weights_list[k][j], 4)
            table_data.append(row)

        df = pd.DataFrame(table_data)
        return df

    def create_table_4(self) -> pd.DataFrame:
        """Создание Таблицы 4: Weighted normalized decision matrixes"""
        table_data = []

        for i, alt in enumerate(self.alternatives):
            row = {'Candidate': alt}
            for k, dm in enumerate(self.dms):
                for j, criterion in enumerate(self.criteria):
                    col_name = f"{dm['id']}_{criterion['name']}"
                    row[col_name] = round(self.Y_list[k][i, j], 4)
            table_data.append(row)

        df = pd.DataFrame(table_data)
        return df

    def create_table_5(self) -> pd.DataFrame:
        """Создание Таблицы 5: Ideal solutions"""
        table_data = []

        Y_star = np.array(self.results['Y_star'])
        Y_l = np.array(self.results['Y_l'])
        Y_r = np.array(self.results['Y_r'])

        for i, alt in enumerate(self.alternatives):
            row = {'Candidate': alt}

            # PIS
            for j, criterion in enumerate(self.criteria):
                col_name = f"PIS_{criterion['name']}"
                row[col_name] = round(Y_star[i, j], 4)

            # L-NIS
            for j, criterion in enumerate(self.criteria):
                col_name = f"L-NIS_{criterion['name']}"
                row[col_name] = round(Y_l[i, j], 4)

            # R-NIS
            for j, criterion in enumerate(self.criteria):
                col_name = f"R-NIS_{criterion['name']}"
                row[col_name] = round(Y_r[i, j], 4)

            table_data.append(row)

        df = pd.DataFrame(table_data)
        return df

    def create_table_6(self) -> pd.DataFrame:
        """Создание Таблицы 6: Separations, closeness coefficients, weights and ranking"""
        table_data = []

        S_plus = self.results['S_plus']
        S_l_minus = self.results['S_l_minus']
        S_r_minus = self.results['S_r_minus']
        C = self.results['C']
        lambda_weights = self.results['lambda_weights']

        for k in range(self.t):
            row = {
                'DM': self.dm_ids[k],
                'S_k+': round(S_plus[k], 4),
                'S_l-': round(S_l_minus[k], 4),
                'S_r-': round(S_r_minus[k], 4),
                'C_k': round(C[k], 4),
                'λ_k': round(lambda_weights[k], 4)
            }
            table_data.append(row)

        df = pd.DataFrame(table_data)
        # Добавляем ранги
        df['Rank'] = df['λ_k'].rank(ascending=False, method='dense').astype(int)
        df = df.sort_values('Rank')
        return df

    def create_table_7(self) -> pd.DataFrame:
        """Создание Таблицы 7: Итоговое групповое решение с суммой по каждому критерию"""
        table_data = []

        Y_group = np.array(self.results['Y_group'])
        ranking_data = self.results['alternatives_ranking']

        # Создаем словарь для рангов
        rank_dict = {item['Alternative']: item['Rank'] for item in ranking_data}

        # Получаем количество критериев
        num_criteria = len(self.criteria)

        for i, alt in enumerate(self.alternatives):
            row = {'Candidate': alt}

            # Для каждого критерия добавляем столбец с суммой
            total_score = 0
            for j in range(num_criteria):
                criterion_name = self.criteria[j]['name']
                col_name = f"Sum_{criterion_name}"
                value = round(Y_group[i, j], 4)
                row[col_name] = value
                total_score += value

            # Общая сумма по всем критериям
            row['Total_Score'] = round(total_score, 4)

            table_data.append(row)

        df = pd.DataFrame(table_data)

        # Добавляем ранги
        df['Rank'] = df['Candidate'].map(rank_dict)
        df = df.sort_values('Rank')

        return df

    def save_all_tables_to_excel(self, filename: str = 'all_tables.xlsx'):
        """Сохранение всех таблиц в Excel файл с названиями"""
        # Создаем все таблицы
        tables = {
            'table_1': self.create_table_1(),
            'table_2': self.create_table_2(),
            'table_3': self.create_table_3(),
            'table_4': self.create_table_4(),
            'table_5': self.create_table_5(),
            'table_6': self.create_table_6(),
            'table_7': self.create_table_7()
        }

        # Названия таблиц
        table_names = {
            'table_1': "ТАБЛИЦА 1: Decision matrixes of example-subjective attributes",
            'table_2': "ТАБЛИЦА 2: Normalized decision matrixes",
            'table_3': "ТАБЛИЦА 3: Weights on attributes of example",
            'table_4': "ТАБЛИЦА 4: Weighted normalized decision matrixes",
            'table_5': "ТАБЛИЦА 5: Ideal solutions",
            'table_6': "ТАБЛИЦА 6: Separations, closeness coefficients, weights and ranking",
            'table_7': "ТАБЛИЦА 7: Integrated assessment of candidates"
        }

        # Сохраняем в Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            workbook = writer.book
            worksheet = workbook.create_sheet('All_Tables')
            writer.sheets['All_Tables'] = worksheet

            start_row = 1

            for i in range(1, 8):
                table_name = f'table_{i}'
                if table_name in tables:
                    # Записываем название таблицы
                    title_cell = worksheet.cell(row=start_row, column=1)
                    title_cell.value = table_names[table_name]
                    title_cell.font = Font(bold=True)
                    title_cell.alignment = Alignment(horizontal='left')

                    start_row += 1

                    # Записываем таблицу
                    df = tables[table_name]

                    # Записываем заголовки
                    for col_idx, col_name in enumerate(df.columns, start=1):
                        header_cell = worksheet.cell(row=start_row, column=col_idx)
                        header_cell.value = col_name
                        header_cell.font = Font(bold=True)
                        header_cell.alignment = Alignment(horizontal='left')

                    start_row += 1

                    # Записываем данные
                    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row):
                        for col_idx, value in enumerate(row, start=1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.value = value
                            cell.alignment = Alignment(horizontal='left')

                    start_row += len(df) + 3  # Пропускаем 3 строки после таблицы

                    print(f"Таблица {i} сохранена: {table_names[table_name]}")

        print(f"\nВсе таблицы сохранены в файл: {filename}")
        return tables

    def print_all_tables(self):
        """Печать всех таблиц в консоль"""
        print("\n" + "="*80)
        print("ПЕЧАТЬ ВСЕХ ТАБЛИЦ:")
        print("="*80)

        tables = {
            'table_1': self.create_table_1(),
            'table_2': self.create_table_2(),
            'table_3': self.create_table_3(),
            'table_4': self.create_table_4(),
            'table_5': self.create_table_5(),
            'table_6': self.create_table_6(),
            'table_7': self.create_table_7()
        }

        table_names = {
            'table_1': "ТАБЛИЦА 1: Decision matrixes",
            'table_2': "ТАБЛИЦА 2: Normalized decision matrixes",
            'table_3': "ТАБЛИЦА 3: Weights on attributes",
            'table_4': "ТАБЛИЦА 4: Weighted normalized decision matrixes",
            'table_5': "ТАБЛИЦА 5: Ideal solutions",
            'table_6': "ТАБЛИЦА 6: Separations, closeness coefficients, weights and ranking",
            'table_7': "ТАБЛИЦА 7: Integrated assessment of candidates"
        }

        for i in range(1, 8):
            table_name = f'table_{i}'
            if table_name in tables:
                print(f"\n--- {table_names[table_name]} ---")
                print(tables[table_name].to_string(index=False))
